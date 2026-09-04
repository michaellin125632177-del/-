# -*- coding: utf-8 -*-
"""日三牙醫體系 統一出勤表
分頁:說明 / 設定 / 醫師班表 / 醫師月結 / 助理班表 / 打卡匯入 / 出勤紀錄 / 月結統計

架構重點:全體系一個檔。
  醫師走「診次制」——一位醫師一列,橫向 31 天 × 早/午/晚,格內填院所代碼。
  一格只容得下一間院所,同一診次被排到兩間院所在結構上就不可能發生。
  助理與醫護長走「工時制」——沿用班表 + 打卡 + 逐日出勤紀錄那一套。
"""
import datetime as dt
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from weekly import W as WEEKLY_RAW, SESSION_TIME, NOTES
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

F = "微軟正黑體"
OUT = "/home/user/-/out/日三牙醫體系_統一出勤表.xlsx"

YEAR, MONTH = 2026, 10
DAYS_IN_MONTH = 31
SESSIONS = ["早", "午", "晚"]

N_DOC   = 40      # 醫師班表列數(24 位 + 保留)
N_ASST_SLOTS = 60
N_ASST  = N_ASST_SLOTS  # 助理/醫護長列數(待名單確認後調整)
PUNCH_N = 2000    # 打卡匯入資料列數

# ---------------------------------------------------------------- 樣式
def font(sz=10, b=False, color="000000"):
    return Font(name=F, size=sz, bold=b, color=color)

TITLE_F   = font(16, True)
HDR_F     = font(10, True, "FFFFFF")
HDR_FILL  = PatternFill("solid", fgColor="1F4E5F")
SUB_FILL  = PatternFill("solid", fgColor="DCE6EC")
IN_FILL   = PatternFill("solid", fgColor="FFF2CC")
CALC_FILL = PatternFill("solid", fgColor="F2F2F2")
WKND_FILL = PatternFill("solid", fgColor="DDEBF7")
LEAVE_FILL= PatternFill("solid", fgColor="D9D9D9")
ALERT_FILL= PatternFill("solid", fgColor="FF9999")
GAP_FILL  = PatternFill("solid", fgColor="FCE4E4")
OT_FILL   = PatternFill("solid", fgColor="FCE4D6")
# 五間院所各自的底色,讓醫師整月動線一眼看得出來
CLINIC_FILL = {
    "悅": PatternFill("solid", fgColor="D9E7F5"),
    "睿": PatternFill("solid", fgColor="DCEEDC"),
    "匯": PatternFill("solid", fgColor="FBE6D4"),
    "曜": PatternFill("solid", fgColor="E8DFF2"),
    "寶": PatternFill("solid", fgColor="FBE3EC"),
}

thin = Side(style="thin", color="AAAAAA")
BOX  = Border(left=thin, right=thin, top=thin, bottom=thin)
DAYSEP = Border(left=Side(style="medium", color="7F7F7F"),
                right=thin, top=thin, bottom=thin)
CTR  = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

FMT_TIME = "hh:mm"; FMT_HR = '0.00;-0.00;"–"'
FMT_MIN  = '0;-0;"–"'; FMT_DATE = "yyyy/m/d"; FMT_CNT = '0;-0;"–"'

def put(ws, cell, value, f=None, fill=None, align=None, fmt=None, border=True):
    c = ws[cell]; c.value = value
    c.font = f or font()
    if fill: c.fill = fill
    if align: c.alignment = align
    if fmt: c.number_format = fmt
    if border: c.border = BOX
    return c

def header_row(ws, row, labels, start_col=2, height=None):
    for i, lab in enumerate(labels):
        c = ws.cell(row=row, column=start_col + i, value=lab)
        c.font, c.fill, c.alignment, c.border = HDR_F, HDR_FILL, CTR, BOX
    if height: ws.row_dimensions[row].height = height

T = lambda h, m: dt.time(h, m)

# ================================================================ 主資料
# 院所:代碼, 簡稱, 全名, 院長, 電話, 地址
CLINICS = [
    ("悅", "晶悅",   "晶悅牙醫診所",   "林智俊", "03-4257710", "320桃園市中壢區新生路251號"),
    ("睿", "晶睿",   "晶睿牙醫診所",   "温健成", "03-4522360", "桃園市中壢區福州二街502號"),
    ("匯", "晶匯",   "晶匯牙醫診所",   "陳昺元", "03-4550719", "桃園市中壢區成章一街172號"),
    ("曜", "晶曜",   "晶曜牙醫診所",   "劉立德", "03-3150666", "桃園市桃園區中正路1398-1號"),
    ("寶", "寶貝牙", "寶貝牙牙醫診所", "黃育亭", "03-2871688", "桃園市大園區新生路四段273號2樓"),
]
CLINIC_CODES = [c[0] for c in CLINICS]
SHORT2CODE   = {c[1]: c[0] for c in CLINICS}

# 醫師:編號, 姓名, 英文名, 專科, 院長職務, 服務院所(簡稱)
DOCTORS = [
 ("D001","林智俊","JASON LIN","牙周植牙專科","晶悅院長",["晶悅","晶睿","晶匯","晶曜"]),
 ("D002","劉立德","LEADER LIU","植牙專科 家庭牙醫專科","晶曜院長",["晶睿","晶曜"]),
 ("D003","王泳泉","RYAN WANG","矯正專科","",["晶悅","晶睿"]),
 ("D004","劉玠旻","CHIEH-MIN LIU","兒童專科","",["晶悅","晶睿","晶曜"]),
 ("D005","陳志僑","CHIH-CHIAO CHEN","口外專科","",["晶悅","晶睿"]),
 ("D006","黃榆鈞","YU-CHUN HUANG","矯正專科","",["晶悅","寶貝牙"]),
 ("D007","温翊君","YI-CHUN WEN","根管治療專科","",["晶悅"]),
 ("D008","林紫亭","TZU-TING LIN","","",["晶悅"]),
 ("D009","陳昺元","PING-YUAN CHEN","美學植牙專科","晶匯院長",["晶悅","晶匯","寶貝牙"]),
 ("D010","蔡季軒","CHI-HSUAN TSAI","","",["晶悅"]),
 ("D011","温健成","JIAN-CHENG WEN","","晶睿院長",["晶睿","晶匯"]),
 ("D012","陳立軒","LI-HSUAN CHEN","","",["晶睿","晶匯","晶曜"]),
 ("D013","朱柏非","PO-FEI CHU","","",["晶悅","晶匯","寶貝牙"]),
 ("D014","鄭漢賢","HON-YIN CHENG","","",["晶悅","晶匯"]),
 ("D015","徐平","PING HSU","","",["晶悅","晶曜"]),
 ("D016","賴敏傑","MIN-CHIEH LAI","","",["晶睿"]),
 ("D017","吳柏賢","BO-XIAN WU","","",["晶睿","晶曜"]),
 ("D018","許博諺","PO-YEN HSU","","",["晶睿","晶曜"]),
 ("D019","翁崎紘","CHI-HONG WENG","","",["晶悅"]),
 ("D020","余柏萱","PO-HSUAN YU","","",["晶睿","寶貝牙"]),
 ("D021","鄭博文","PO-WEN CHENG","","",["晶悅"]),
 ("D022","黃育亭","YU-TING HUANG","","寶貝牙院長",["寶貝牙"]),
 ("D023","黃婷愉","TING-YU HUANG","","",["寶貝牙"]),
 ("D024","林紓安","SHU-AN LIN","","",["晶匯","寶貝牙"]),
 # 以下三位只出現在門診表,官網醫療團隊頁的 24 位名單中沒有,英文名與專科待補
 ("D025","陳爵安","","","",["晶睿"]),
 ("D026","吳冠廷","","","",["晶匯"]),
 ("D027","楊孟庭","","","",["晶匯"]),
]
# 隔週互換的四組:同一位醫師同一時段在兩間院所輪替。
# 預設在「第 1、3、5 個該星期幾」出現;列在這裡的改為第 2、4 個。
# ⚠ 哪一週在哪一間目前無從得知,這是暫定假設,需院所確認。
ALT_EVEN = {("劉立德",3,"曜"), ("朱柏非",6,"匯"), ("王泳泉",1,"睿"), ("陳昺元",6,"匯")}

# 助理與醫護長名單:待提供
ASSISTANTS = []      # (編號, 姓名, 職類, 院所簡稱, 到職日, 備註)

# 助理/醫護長 班別代碼:代碼,名稱,應到,應退,休息分,排班工時,計出勤,類別
WORK_CODES = [
    ("A",  "早班",          T(9,0),  T(18,0),  60, 8.0, 1, "上班"),
    ("P",  "晚班",          T(12,30),T(21,30), 60, 8.0, 1, "上班"),
    ("MD", "中班",          T(11,0), T(20,0),  60, 8.0, 1, "上班"),
    ("ADM","行政班",        T(9,0),  T(18,0),  60, 8.0, 1, "上班"),
    ("支", "支援他院",      None,    None,     60, 8.0, 1, "上班"),
    ("訓", "教育訓練/會議",  None,    None,     60, 8.0, 1, "上班"),
    ("OFF","排休",          None,    None,      0, 0.0, 0, "休假"),
    ("特", "特休",          None,    None,      0, 0.0, 0, "休假"),
    ("病", "病假",          None,    None,      0, 0.0, 0, "休假"),
    ("事", "事假",          None,    None,      0, 0.0, 0, "休假"),
    ("公", "公假",          None,    None,      0, 0.0, 0, "休假"),
    ("國", "國定假日",      None,    None,      0, 0.0, 0, "休假"),
    ("休", "休診",          None,    None,      0, 0.0, 0, "休假"),
]
for c in WORK_CODES:                      # 工時定義一致性檢查
    if c[2] and c[3]:
        span = (dt.datetime.combine(dt.date.min, c[3])
                - dt.datetime.combine(dt.date.min, c[2])).seconds / 3600
        assert abs(span - c[4]/60 - c[5]) < 1e-9, f"代碼 {c[0]} 工時定義不一致"

DOC_CODES = ([(c[0], f"{c[1]}牙醫") for c in CLINICS] +
             [("訓", "教育訓練 / 學會"), ("OFF", "排休"), ("特", "特休"),
              ("病", "病假"), ("事", "事假"), ("公", "公假"),
              ("國", "國定假日"), ("休", "休診")])
DOC_LEAVE = [c[0] for c in DOC_CODES if c[0] not in CLINIC_CODES and c[0] != "訓"]

# ================================================================ 設定分頁座標
SET_Y, SET_M = "設定!$C$2", "設定!$D$2"
WC_R0 = 6;   WC_R1 = WC_R0 + len(WORK_CODES) - 1        # 6..18
WC_WORK1 = WC_R0 + 5                                     # 上班類最後一列 11
DC_R0 = 23;  DC_R1 = DC_R0 + len(DOC_CODES) - 1          # 23..35
PM_R0 = 39                                               # 參數:40,41,42
CL_R0 = 46;  CL_L0 = 47; CL_L1 = CL_L0 + 6               # 47..53
PP_R0 = 56;  PP_L0 = 57; PP_L1 = PP_L0 + 139             # 57..196

R_WC     = f"設定!$B${WC_R0}:$B${WC_R1}"
R_WC_W   = f"設定!$B${WC_R0}:$B${WC_WORK1}"
R_WC_L   = f"設定!$B${WC_WORK1+1}:$B${WC_R1}"
R_WC_IN  = f"設定!$D${WC_R0}:$D${WC_R1}"
R_WC_OUT = f"設定!$E${WC_R0}:$E${WC_R1}"
R_WC_RST = f"設定!$F${WC_R0}:$F${WC_R1}"
R_WC_HRS = f"設定!$G${WC_R0}:$G${WC_R1}"
R_WC_ATT = f"設定!$H${WC_R0}:$H${WC_R1}"
R_DC     = f"設定!$B${DC_R0}:$B${DC_R1}"
P_OT_MIN, P_OT_UNIT, P_GRACE = "設定!$C$40", "設定!$C$41", "設定!$C$42"
R_EID  = f"設定!$B${PP_L0}:$B${PP_L1}"
R_NAME = f"設定!$C${PP_L0}:$C${PP_L1}"
R_ROLE = f"設定!$D${PP_L0}:$D${PP_L1}"
R_SPEC = f"設定!$E${PP_L0}:$E${PP_L1}"
R_DUTY = f"設定!$F${PP_L0}:$F${PP_L1}"
R_HOME = f"設定!$G${PP_L0}:$G${PP_L1}"
R_WEEK = "設定!$O$6:$O$12"
DAYS_FX = f"DAY(EOMONTH(DATE({SET_Y},{SET_M},1),0))"

# ── 門診表 → 每位醫師的每週時段 ────────────────────────────
DOC_WEEK = {}                       # 姓名 -> {(星期, 診次): (院所代碼, 標記)}
for _cl, _cells in WEEKLY_RAW.items():
    for (_wd, _ss), _lst in _cells.items():
        for _nm, _flag in _lst:
            DOC_WEEK.setdefault(_nm, {})[(_wd, _ss, _cl)] = _flag
NAME2EID = {d[1]: d[0] for d in DOCTORS}
# 各院所有排診的時段 = 有開診。週日五院全休;週六晚診五院全休;寶貝牙平日早診休診。
OPEN = {c[0]: {(w, t) for (w, t) in WEEKLY_RAW.get(c[0], {})} for c in CLINICS}
ALL_CLOSED = {(w, t) for w in range(1, 8) for t in range(3)
              if not any((w, t) in OPEN[c] for c in OPEN)}
_unknown = [n for n in DOC_WEEK if n not in NAME2EID]
assert not _unknown, f"門診表出現名冊沒有的醫師:{_unknown}"

def expand_month(name, year, month, ndays):
    """把每週固定門診表展開成當月 31 天 × 3 診次。"""
    out = []
    slots = DOC_WEEK.get(name, {})
    for d in range(1, ndays + 1):
        date = dt.date(year, month, d)
        wd = date.weekday() + 1                      # 1=一 … 7=日
        occ = (d - 1) // 7 + 1                       # 當月第幾個該星期幾
        for ss in range(3):
            if (wd, ss) in ALL_CLOSED:      # 五間院所該時段全休
                out.append("休"); continue
            hit = [(cl, fg) for (w, t, cl), fg in slots.items() if w == wd and t == ss]
            if not hit:
                out.append(""); continue
            val = ""
            for cl, fg in hit:
                if "~" in fg:                        # 隔週
                    want_odd = (name, wd, cl) not in ALT_EVEN
                    if (occ % 2 == 1) != want_odd:
                        continue
                val = cl
                break
            out.append(val)
    return out

wb = Workbook()

# ================================================================ 1. 說明
ws = wb.active; ws.title = "說明"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 106
put(ws, "B2", "日三牙醫體系 — 統一出勤表 使用說明", TITLE_F, border=False)
ws.merge_cells("B2:C2")

BLOCKS = [
 ("為什麼是一個檔",
  "體系 27 位醫師裡有 18 位跨院看診,林智俊院長一個人就跑四間。\n"
  "若一間院所一個檔,排他一個月要開四個檔,而且沒有任何機制擋得住「同一個早診\n"
  "同時被排在晶悅和晶睿」。所以全體系共用這一個檔。"),
 ("醫師與助理為什麼分開排",
  "醫師的一天是「幾個診次」,助理的一天是「幾個小時」,兩者不能塞進同一種格子。\n"
  "醫師 → 醫師班表:一人一列,橫向 31 天 × 早/午/晚,格內填院所代碼(悅睿匯曜寶)。\n"
  "   一格只容得下一間院所,衝堂在結構上就不可能發生。\n"
  "助理與醫護長 → 助理班表:一人一列,橫向 31 天,格內填班別代碼(A/P/OFF/特…)。"),
 ("九個分頁怎麼分工",
  "① 說明 ② 設定 — 總部維護,各院所勿動。\n"
  "③ 醫師週班表 — 官網門診表的固定週輪值,是排月班的母表。\n"
  "④ 醫師班表 ⑤ 醫師月結 — 本月實際診次,已依週班表填好。\n"
  "⑥ 助理班表 — 各院所醫護長填自己院所那幾列。\n"
  "⑦ 打卡匯入 ⑧ 出勤紀錄 ⑨ 月結統計 — 助理與醫護長的法定出勤與月結。"),
 ("週班表與月班表的關係",
  "門診表是固定的每週輪值,所以月班表不必一格一格填——「醫師班表」已經照\n"
  "「醫師週班表」把整個月展開好了,直接改例外即可(請假就把該格改成假別代碼)。\n"
  "⚠ 改「設定」的本期年月只會更新日期與星期列,格子裡的班不會重排。換月份時\n"
  "   請告知重新產生,或自行把上個月的格子整段調整。"),
 ("各院所診次時間不一樣",
  "晶悅/晶睿/晶匯/晶曜:早 09:00~11:30、午 14:00~16:30、晚 18:00~20:30。\n"
  "寶貝牙:早 09:00~14:00、午 14:00~17:00、晚 17:30~21:00,且週一至週五早診休診。\n"
  "五間院所週日全休、週六晚診全休。醫師班表的早/午/晚只是時段代號,\n"
  "實際幾點到幾點依「設定」院所清單裡該院所的定義。開診時段表也在那裡。"),
 ("每月作業流程",
  "上月底:總部排醫師班表 → 看底下五列「各院所每診次醫師數」有沒有 0 → 發給各院所。\n"
  "     各院所醫護長排助理班表 → 檢查人力檢核列 → 交回。\n"
  "當月底:打卡機匯出檔貼進「打卡匯入」→ 到「出勤紀錄」處理異常 → 月結送人資。"),
 ("醫師班表怎麼填",
  "每一天有三格:早診、午診、晚診。格內填院所代碼,空白代表該診次沒排班。\n"
  "  悅 = 晶悅   睿 = 晶睿   匯 = 晶匯   曜 = 晶曜   寶 = 寶貝牙\n"
  "請假也是填在診次格裡(特/病/事/公/國/OFF/休),因為醫師請假常常只請半天。\n"
  "所以醫師的月結一律以「診次」為單位,不是天數——請假 3 個診次就是請一天。\n"
  "五種院所代碼各有底色,一位醫師整月在五間之間怎麼跑,橫著看一列就知道。"),
 ("出勤紀錄怎麼看",
  "一人一天一列,自動把助理班表與打卡對起來,判定五種狀態:\n"
  "正常、遲到、早退、未打卡、假日出勤。紅底 = 需處理;橘底 = 當天有加班。\n"
  "用上方篩選鈕只看非「正常」的最快。沒有打卡機的院所,可把「實際上班/下班」\n"
  "兩欄公式刪掉改成手填,其餘照算。"),
 ("加班怎麼認定",
  "超過排班工時的分鐘數要跨過門檻(預設 30 分)才算加班,並無條件捨去到計算\n"
  "單位(預設 30 分)。否則每天晚幾分鐘打卡都會被算成加班,一個月會憑空多出\n"
  "好幾小時。遲到早退另有寬限(預設 5 分)。三個參數在「設定」的計算參數區。\n"
  "這三個數字是勞資慣例、不是法規,上線前請人資確認,並讓同仁事先知道規則。"),
 ("顏色代表什麼",
  "黃底 = 你要填的格子        灰底 = 公式自動算,別動\n"
  "淺藍欄 = 週六/週日          灰色格 = 休假類代碼\n"
  "五種院所底色 = 醫師該診次在哪一間\n"
  "紅色 = 人力不足或出勤異常    橘色 = 有加班"),
 ("⚠ 法規提醒",
  "依勞動基準法第 30 條,雇主應置備勞工出勤紀錄,逐日記載至分鐘為止,保存 5 年。\n"
  "「出勤紀錄」分頁為此設計,但必須每月另存唯讀封存檔才算數,不要只留一份覆蓋。\n"
  "加班時數為「實際工時 − 排班工時」,未依第 24 條換算費率,薪資須另行計算。\n"
  "受僱醫師是否適用勞基法依僱傭契約與主管機關認定,建議請人資確認後再定義醫師工時。"),
 ("⚠ 需要核對的三件事",
  "一、門診表是從截圖轉錄的,姓名與隔週標記請逐格核對「醫師週班表」。\n"
  "   若能提供五個門診表頁面的網頁存檔,可以完全免除轉錄誤差。\n"
  "二、陳爵安、吳冠廷、楊孟庭三位只出現在門診表,官網醫療團隊頁的名單沒有,\n"
  "   已暫編 D025~D027,英文名與專科待補。\n"
  "三、隔週互換有四組(劉立德週三、朱柏非週六、王泳泉週一、陳昺元週六),\n"
  "   同一時段在兩間院所輪替。目前假設第 1、3、5 個該星期幾在前一間,\n"
  "   第 2、4 個在後一間——哪一週在哪一間無從得知,務必確認。"),
 ("目前待補的資料",
  "一、助理與醫護長名單尚未提供,助理班表/打卡匯入/出勤紀錄/月結統計四個分頁\n"
  "   已經建好結構但沒有人員資料,名單一到即可直接灌入。\n"
  "二、助理的班別時間(早班/晚班/中班/行政班)目前是暫定值,需要確認。\n"
  "三、員工編號目前為暫編 D001~D027,若人事或打卡系統另有編號應以那套為準。\n"
  "四、官網門診表有星號註記者(並非每週固定)需逐一向院所確認。"),
]
r = 4
for h, body in BLOCKS:
    put(ws, f"B{r}", h, font(10, True), SUB_FILL, LEFT)
    put(ws, f"C{r}", body, font(10), None, WRAP)
    ws.row_dimensions[r].height = 15 * (body.count("\n") + 1) + 8
    r += 1

# ================================================================ 2. 設定
st = wb.create_sheet("設定")
st.sheet_view.showGridLines = False
for col, w in {"A":2,"B":11,"C":18,"D":9,"E":9,"F":10,"G":11,"H":8,"I":9,"J":9,
               "K":2,"L":12,"M":2,"N":9,"O":8,"P":2,"Q":13,"R":2,"S":16}.items():
    st.column_dimensions[col].width = w
put(st, "B1", "設定表(總部維護,各院所請勿修改)", TITLE_F, border=False)
put(st, "B2", "本期年月", font(10, True), SUB_FILL, CTR)
put(st, "C2", YEAR, font(10, True), IN_FILL, CTR, "0")
put(st, "D2", MONTH, font(10, True), IN_FILL, CTR, "0")
put(st, "E2", "← 整份檔案的日期、星期都以這裡為準", font(9, color="808080"),
    None, LEFT, border=False)

put(st, f"B{WC_R0-2}", "一、助理/醫護長 班別代碼(工時制)", font(11, True), border=False)
header_row(st, WC_R0-1, ["代碼","名稱","應到","應退","休息(分)","排班工時","計出勤","類別"])
for i, row in enumerate(WORK_CODES):
    for j, v in enumerate(row):
        c = st.cell(row=WC_R0+i, column=2+j, value=v)
        c.font, c.border, c.alignment = font(), BOX, CTR
        if j == 1: c.alignment = LEFT
        if j in (2,3): c.number_format = FMT_TIME
        if j == 5: c.number_format = "0.0"
put(st, f"B{WC_R1+1}",
    "※ 有固定時段的代碼必須滿足「應退 − 應到 − 休息 = 排班工時」,否則出勤紀錄會算出不存在的加班。"
    "建置腳本會自動檢查。", font(9, color="808080"), None, LEFT, border=False)

put(st, f"B{DC_R0-2}", "二、醫師診次代碼", font(11, True), border=False)
header_row(st, DC_R0-1, ["代碼","意義"])
for i, (code, mean) in enumerate(DOC_CODES):
    put(st, f"B{DC_R0+i}", code, font(10, True),
        CLINIC_FILL.get(code, CALC_FILL), CTR)
    put(st, f"C{DC_R0+i}", mean, font(), None, LEFT)
put(st, f"B{DC_R1+1}",
    "※ 醫師班表一格 = 一個診次。填院所代碼表示該診次在哪間看診,填休假代碼表示該診次請假。",
    font(9, color="808080"), None, LEFT, border=False)

put(st, "S5", "助理/醫護長編號(下拉用)", font(9, True), SUB_FILL, CTR)
AST_EIDS = [a[0] for a in ASSISTANTS]
for i in range(N_ASST_SLOTS):
    put(st, f"S{6+i}", AST_EIDS[i] if i < len(AST_EIDS) else "",
        font(9), CALC_FILL if i < len(AST_EIDS) else None, CTR)

put(st, "Q5", "醫師編號(下拉用)", font(9, True), SUB_FILL, CTR)
DOC_EIDS = [d[0] for d in DOCTORS]
for i in range(40):
    put(st, f"Q{6+i}", DOC_EIDS[i] if i < len(DOC_EIDS) else "",
        font(9), CALC_FILL if i < len(DOC_EIDS) else None, CTR)

put(st, "N5", "星期對照", font(9, True), SUB_FILL, CTR)
put(st, "O5", "", font(9, True), SUB_FILL, CTR)
for i, w in enumerate(["日","一","二","三","四","五","六"]):
    put(st, f"N{6+i}", i+1, font(9), CALC_FILL, CTR)
    put(st, f"O{6+i}", w, font(9), CALC_FILL, CTR)

put(st, f"B{PM_R0}", "三、計算參數(全體系一致)", font(11, True), border=False)
header_row(st, PM_R0, ["參數","值","說明"], start_col=2)
PARAMS = [("加班認定門檻(分)", 30, "超過排班工時多少分鐘才認定為加班。低於門檻視為正常收尾。"),
          ("加班計算單位(分)", 30, "認定為加班後,無條件捨去到此單位。填 1 表示逐分鐘計。"),
          ("遲到早退寬限(分)", 5,  "打卡與應到/應退時間差在此範圍內不判定為遲到或早退。")]
for i, (nm, val, desc) in enumerate(PARAMS):
    r = PM_R0 + 1 + i
    put(st, f"B{r}", nm, font(), SUB_FILL, LEFT)
    put(st, f"C{r}", val, font(10, True), IN_FILL, CTR, "0")
    put(st, f"D{r}", desc, font(9), None, LEFT); st.merge_cells(f"D{r}:J{r}")
put(st, f"B{PM_R0+4}",
    "※ 這三個是勞資雙方的認定慣例,不是法律規定的數字。上線前請與人資或勞務顧問確認,"
    "並讓同仁事先知道規則。", font(9, color="808080"), None, LEFT, border=False)

put(st, f"B{CL_R0-1}", "四、院所清單", font(11, True), border=False)
header_row(st, CL_R0, ["代碼","簡稱","全名","院長","電話","地址",
                       "早診起","早診訖","午診起","午診訖","晚診起","晚診訖"])
CLINIC_ROWS = [tuple(c) + tuple(t for pair in SESSION_TIME[c[0]] for t in pair)
               for c in CLINICS]
for i in range(CL_L1 - CL_L0 + 1):
    vals = CLINIC_ROWS[i] if i < len(CLINIC_ROWS) else ("",)*12
    for j, v in enumerate(vals):
        c = st.cell(row=CL_L0+i, column=2+j, value=v)
        c.font, c.border, c.alignment = font(), BOX, CTR
        if j in (2,5): c.alignment = LEFT
        if j >= 6: c.font = font(9)
        if j == 0 and v: c.fill = CLINIC_FILL[v]

put(st, "U{}".format(CL_R0), "開診時段(1 = 有開診,0 = 休診;順序為週一早…週六晚)",
    font(9, True), SUB_FILL, LEFT)
st.merge_cells(start_row=CL_R0, start_column=21, end_row=CL_R0, end_column=38)
for i in range(CL_L1 - CL_L0 + 1):
    code = CLINICS[i][0] if i < len(CLINICS) else None
    for w in range(6):
        for t in range(3):
            c = st.cell(row=CL_L0 + i, column=21 + w * 3 + t)
            c.value = (1 if code and (w + 1, t) in OPEN[code] else 0) if code else None
            c.font, c.border, c.alignment = font(8), BOX, CTR
            if code and (w + 1, t) not in OPEN[code]: c.fill = LEAVE_FILL
for w in range(6):
    for t in range(3):
        c = st.cell(row=CL_R0 - 1, column=21 + w * 3 + t,
                    value=f"{'一二三四五六'[w]}{SESSIONS[t]}")
        c.font, c.fill, c.alignment, c.border = font(8, True, "FFFFFF"), HDR_FILL, CTR, BOX

put(st, f"B{CL_L1+2}",
    "※ 寶貝牙的診次時段與其他四院不同,且週一至週五早診休診、週六晚診休診。"
    "醫師班表的早/午/晚只是時段代號,實際幾點到幾點依上表該院所的定義。",
    font(9, color="808080"), None, LEFT, border=False)

put(st, f"B{PP_R0-1}", "五、人員名冊", font(11, True), border=False)
header_row(st, PP_R0, ["員工編號","姓名","職類","專科","職務","主要院所",
                       "服務院所(醫師)","英文名","到職日","備註"])
ROSTER = []
for eid, nm, eng, spec, duty, teams in DOCTORS:
    home = SHORT2CODE[teams[0]] if not duty else SHORT2CODE[duty.replace("院長","")]
    ROSTER.append([eid, nm, "醫師", spec, duty,
                   next(c[1] for c in CLINICS if c[0] == home),
                   " · ".join(teams), eng, "", ""])
for a in ASSISTANTS:
    ROSTER.append([a[0], a[1], a[2], "", "", a[3], "", "", a[4], a[5]])
for i in range(PP_L1 - PP_L0 + 1):
    vals = ROSTER[i] if i < len(ROSTER) else [""]*10
    for j, v in enumerate(vals):
        c = st.cell(row=PP_L0+i, column=2+j, value=v)
        c.font, c.border, c.alignment = font(9), BOX, CTR
        if i < len(ROSTER): c.fill = IN_FILL
        if j in (3, 6, 9): c.alignment = LEFT
print(f"人員名冊:{len(ROSTER)} 人(醫師 {len(DOCTORS)}、助理/醫護長 {len(ASSISTANTS)})")

# ================================================================ 3. 醫師週班表
wk = wb.create_sheet("醫師週班表")
wk.sheet_view.showGridLines = False
WK_ROW0 = 5
WK_ROW1 = WK_ROW0 + N_DOC - 1
WK_C0 = 3                                   # C 欄起,18 格
wk.freeze_panes = "C5"
wk.column_dimensions["A"].width = 9
wk.column_dimensions["B"].width = 10
for i in range(18):
    wk.column_dimensions[get_column_letter(WK_C0 + i)].width = 5.2
wk.column_dimensions[get_column_letter(WK_C0 + 18)].width = 3
wk.column_dimensions[get_column_letter(WK_C0 + 19)].width = 40
put(wk, "A1", "醫師週班表(固定門診表 · 這是排月班的母表)", TITLE_F, border=False)
wk.merge_cells(start_row=1, start_column=1, end_row=1, end_column=WK_C0 + 17)
put(wk, "A2", "資料來源:官網各院所門診表(自截圖轉錄,請核對)。改這裡不會自動改「醫師班表」,"
    "月班表是照這張表產生出來的固定值。", font(9, color="808080"), None, LEFT, border=False)
for lab, col in (("員工編號", 1), ("姓名", 2)):
    c = wk.cell(row=3, column=col, value=lab)
    c.font, c.fill, c.alignment, c.border = HDR_F, HDR_FILL, CTR, BOX
    wk.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)
WD_LAB = ["一", "二", "三", "四", "五", "六"]
for w in range(6):
    a = WK_C0 + w * 3
    c = wk.cell(row=3, column=a, value=f"週{WD_LAB[w]}")
    c.font, c.fill, c.alignment, c.border = HDR_F, HDR_FILL, CTR, BOX
    wk.merge_cells(start_row=3, start_column=a, end_row=3, end_column=a + 2)
    for sidx in range(3):
        cc = wk.cell(row=4, column=a + sidx, value=SESSIONS[sidx])
        cc.font, cc.fill, cc.alignment = font(8), SUB_FILL, CTR
        cc.border = DAYSEP if sidx == 0 else BOX
put(wk, f"{get_column_letter(WK_C0+19)}3", "附註", HDR_F, HDR_FILL, CTR)
wk.merge_cells(start_row=3, start_column=WK_C0+19, end_row=4, end_column=WK_C0+19)

for i in range(N_DOC):
    r = WK_ROW0 + i
    wk.row_dimensions[r].height = 17
    eid = DOCTORS[i][0] if i < len(DOCTORS) else None
    nm = DOCTORS[i][1] if i < len(DOCTORS) else None
    a = wk.cell(row=r, column=1, value=eid)
    a.font, a.fill, a.border, a.alignment = font(9), IN_FILL, BOX, CTR
    b = wk.cell(row=r, column=2)
    b.value = f'=IFERROR(INDEX({R_NAME},MATCH($A{r},{R_EID},0)),"")'
    b.font, b.fill, b.border, b.alignment = font(9), CALC_FILL, BOX, CTR
    slots = DOC_WEEK.get(nm, {}) if nm else {}
    notes = []
    for w in range(6):
        for sidx in range(3):
            cc = wk.cell(row=r, column=WK_C0 + w * 3 + sidx)
            hit = [(cl, fg) for (ww, tt, cl), fg in slots.items()
                   if ww == w + 1 and tt == sidx]
            if hit:
                cc.value = "".join(f"{cl}{fg}" for cl, fg in sorted(hit))
                cc.fill = CLINIC_FILL.get(hit[0][0], CALC_FILL)
            cc.font, cc.alignment = font(8, True), CTR
            cc.border = DAYSEP if sidx == 0 else BOX
    for (ww, tt, cl), fg in sorted(slots.items()):
        if "!" in fg and (cl, ww, tt) not in ():
            key = (cl, ww, tt)
        if "!" in fg:
            notes.append(NOTES.get((cl, ww, tt), "有附加條件"))
    n = wk.cell(row=r, column=WK_C0 + 19, value=" / ".join(dict.fromkeys(notes)))
    n.font, n.border, n.alignment = font(8), BOX, LEFT
NOTE0 = WK_ROW1 + 2
for i, t in enumerate([
  "※ 標記:~ = 隔週看診   * = 官網註明並非每週固定,詳情請聯繫院所   格內兩個代碼 = 兩間院所輪替。",
  "※ 隔週的四組互換(劉立德週三、朱柏非週六、王泳泉週一、陳昺元週六)目前假設在「第 1、3、5 個該星期幾」"
  "出現於前一間、第 2、4 個出現於後一間。哪一週在哪一間需要院所確認。",
  "※ 週日五間院所皆休診。寶貝牙另有週一至週五早診休診、週六晚診休診。",
]):
    put(wk, f"A{NOTE0+i}", t, font(9, color="808080"), None, LEFT, border=False)
dv_wk = DataValidation(type="list", formula1="設定!$Q$6:$Q$45", allow_blank=True)
wk.add_data_validation(dv_wk); dv_wk.add(f"A{WK_ROW0}:A{WK_ROW1}")

# ================================================================ 4. 醫師班表
ds = wb.create_sheet("醫師班表")
ds.sheet_view.showGridLines = False
DS_C0 = 5                                   # E 欄起
DS_ROW0 = 6
DS_ROW1 = DS_ROW0 + N_DOC - 1               # 45
ds.freeze_panes = "E6"
for col, w in {"A":9,"B":10,"C":15,"D":11}.items():
    ds.column_dimensions[col].width = w
def dcol(d, sidx):                           # 第 d 日、第 sidx 診次的欄號
    return DS_C0 + (d-1)*3 + sidx
for d in range(1, DAYS_IN_MONTH+1):
    for sidx in range(3):
        ds.column_dimensions[get_column_letter(dcol(d, sidx))].width = 3.4
LAST_C = dcol(DAYS_IN_MONTH, 2)

put(ds, "A1", "醫師班表(診次制 · 一格 = 一個診次 · 已依週班表填好本月)", TITLE_F, border=False)
ds.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(LAST_C, 40))
put(ds, "A2", "期間", font(10, True), SUB_FILL, CTR)
ds["B2"] = f'={SET_Y}&"年"&{SET_M}&"月"'
ds["B2"].font, ds["B2"].fill, ds["B2"].alignment, ds["B2"].border = (
    font(10, True), CALC_FILL, CTR, BOX)
put(ds, "D2", "圖例:", font(9, True), None, LEFT, border=False)
for i, (code, short, *_rest) in enumerate(CLINICS):
    c = ds.cell(row=2, column=DS_C0 + i*4)
    c.value = f"{code} {short}"
    c.font, c.fill, c.alignment, c.border = font(9), CLINIC_FILL[code], CTR, BOX
    ds.merge_cells(start_row=2, start_column=DS_C0+i*4, end_row=2, end_column=DS_C0+i*4+3)

for lab, col in (("員工編號",1), ("姓名",2), ("專科",3), ("職務",4)):
    c = ds.cell(row=3, column=col, value=lab)
    c.font, c.fill, c.alignment, c.border = HDR_F, HDR_FILL, CTR, BOX
    ds.merge_cells(start_row=3, start_column=col, end_row=5, end_column=col)

for d in range(1, DAYS_IN_MONTH+1):
    a = dcol(d, 0); L = get_column_letter(a)
    c = ds.cell(row=3, column=a, value=f'=IF({d}>{DAYS_FX},"",{d})')
    c.font, c.fill, c.alignment, c.border = HDR_F, HDR_FILL, CTR, BOX
    ds.merge_cells(start_row=3, start_column=a, end_row=3, end_column=a+2)
    for sidx in range(3):
        cc = dcol(d, sidx); CL = get_column_letter(cc)
        w = ds.cell(row=4, column=cc)
        w.value = (f'=IF({d}>{DAYS_FX},"",'
                   f'INDEX({R_WEEK},WEEKDAY(DATE({SET_Y},{SET_M},{d}),1)))')
        w.font, w.fill, w.alignment = font(8, True), SUB_FILL, CTR
        w.border = DAYSEP if sidx == 0 else BOX
        s = ds.cell(row=5, column=cc, value=SESSIONS[sidx])
        s.font, s.fill, s.alignment = font(8), SUB_FILL, CTR
        s.border = DAYSEP if sidx == 0 else BOX

for i in range(N_DOC):
    r = DS_ROW0 + i
    ds.row_dimensions[r].height = 17
    eid = DOCTORS[i][0] if i < len(DOCTORS) else None
    a = ds.cell(row=r, column=1, value=eid)
    a.font, a.fill, a.border, a.alignment = font(9), IN_FILL, BOX, CTR
    for col, src in ((2, R_NAME), (3, R_SPEC), (4, R_DUTY)):
        c = ds.cell(row=r, column=col)
        c.value = f'=IFERROR(INDEX({src},MATCH($A{r},{R_EID},0)),"")'
        c.font, c.fill, c.border, c.alignment = font(9), CALC_FILL, BOX, CTR
        if col == 3: c.alignment = LEFT
    nm = DOCTORS[i][1] if i < len(DOCTORS) else None
    month_vals = expand_month(nm, YEAR, MONTH, DAYS_IN_MONTH) if nm else None
    for d in range(1, DAYS_IN_MONTH+1):
        for sidx in range(3):
            cc = ds.cell(row=r, column=dcol(d, sidx))
            if month_vals:
                v = month_vals[(d-1)*3 + sidx]
                if v: cc.value = v
            cc.font, cc.alignment = font(9, True), CTR
            cc.border = DAYSEP if sidx == 0 else BOX

TALLY0 = DS_ROW1 + 2                         # 47
put(ds, f"A{TALLY0-1}", "各院所每診次醫師數(自動計算,0 表示該診次沒有醫師)",
    font(10, True), SUB_FILL, LEFT)
ds.merge_cells(start_row=TALLY0-1, start_column=1, end_row=TALLY0-1, end_column=4)
for k, (code, short, *_r) in enumerate(CLINICS):
    r = TALLY0 + k
    put(ds, f"A{r}", f"{code} {short}", font(9, True), CLINIC_FILL[code], CTR)
    ds.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    for d in range(1, DAYS_IN_MONTH+1):
        for sidx in range(3):
            cc = dcol(d, sidx); CL = get_column_letter(cc)
            c = ds.cell(row=r, column=cc)
            orow = CL_L0 + k
            c.value = (f'=IF({CL}$4="","",'
                       f'IF(INDEX(設定!$U${orow}:$AL${orow},'
                       f'(WEEKDAY(DATE({SET_Y},{SET_M},{d}),2)-1)*3+{sidx+1})=0,"休",'
                       f'COUNTIF({CL}${DS_ROW0}:{CL}${DS_ROW1},"{code}")))')
            c.font, c.fill, c.alignment = font(8, True), CALC_FILL, CTR
            c.border = DAYSEP if sidx == 0 else BOX
            c.number_format = FMT_CNT
for _i, _t in enumerate([
  "※ 一格只容得下一間院所,所以同一位醫師同一診次被排到兩間院所在結構上不可能發生。"
  "上面五列是反向檢查:某間院所某個診次掛 0,表示那個時段沒有醫師。",
  "※ 本表已依「醫師週班表」把固定門診展開成本月的值,可以直接改。請假就把該格改成假別代碼。",
  "※ 換月份時:改「設定」的本期年月只會更新上方日期與星期,格子裡的班不會自動重排——"
  "固定班表是照週班表產生的值,需要重新產生。",
], ):
    put(ds, f"A{TALLY0+len(CLINICS)+1+_i}", _t, font(9, color="808080"), None, LEFT, border=False)

dv_doc = DataValidation(type="list", formula1=R_DC, allow_blank=True,
                        showErrorMessage=True, errorTitle="診次代碼無效",
                        error="請填院所代碼(悅睿匯曜寶)或休假代碼。")
ds.add_data_validation(dv_doc)
dv_doc.add(f"{get_column_letter(DS_C0)}{DS_ROW0}:{get_column_letter(LAST_C)}{DS_ROW1}")
dv_deid = DataValidation(type="list", formula1="設定!$Q$6:$Q$45", allow_blank=True)
ds.add_data_validation(dv_deid); dv_deid.add(f"A{DS_ROW0}:A{DS_ROW1}")

E0 = get_column_letter(DS_C0); EL = get_column_letter(LAST_C)
GRID = f"{E0}{DS_ROW0}:{EL}{DS_ROW1}"
for code in CLINIC_CODES:
    ds.conditional_formatting.add(GRID, FormulaRule(
        formula=[f'{E0}{DS_ROW0}="{code}"'], fill=CLINIC_FILL[code], stopIfTrue=True))
ds.conditional_formatting.add(GRID, FormulaRule(
    formula=[f'COUNTIF({R_DC},{E0}{DS_ROW0})>0'], fill=LEAVE_FILL, stopIfTrue=True))
ds.conditional_formatting.add(GRID, FormulaRule(
    formula=[f'OR({E0}$4="六",{E0}$4="日")'], fill=WKND_FILL))
ds.conditional_formatting.add(f"{E0}4:{EL}5", FormulaRule(
    formula=[f'OR({E0}$4="六",{E0}$4="日")'], fill=WKND_FILL))
for k, code in enumerate(CLINIC_CODES):
    r = TALLY0 + k
    ds.conditional_formatting.add(f"{E0}{r}:{EL}{r}", FormulaRule(
        formula=[f'AND({E0}$4<>"",ISNUMBER({E0}{r}),{E0}{r}=0)'], fill=ALERT_FILL))

# ================================================================ 4. 醫師月結
dm = wb.create_sheet("醫師月結")
dm.sheet_view.showGridLines = False
dm.freeze_panes = "C5"
DM_COLS = ([("A","員工編號",11), ("B","姓名",11), ("C","專科",15), ("D","職務",11)]
           + [(get_column_letter(5+i), f"{c[1]}診次", 10) for i, c in enumerate(CLINICS)]
           + [("J","總診次",9), ("K","服務院所數",11), ("L","教育訓練",9),
              ("M","排休",7), ("N","特休",7), ("O","病假",7), ("P","事假",7),
              ("Q","公假",7), ("R","國定假日",9), ("S","休診",7),
              ("T","休假/休診合計",12), ("U","未排診次",10)])
for col, _, w in DM_COLS: dm.column_dimensions[col].width = w
put(dm, "A1", "醫師月結(全自動 · 以診次為單位)", TITLE_F, border=False)
dm.merge_cells("A1:U1")
put(dm, "A2", "期間", font(10, True), SUB_FILL, CTR)
dm["A2"].alignment = CTR
dm["B2"] = f'={SET_Y}&"年"&{SET_M}&"月"'
dm["B2"].font, dm["B2"].fill, dm["B2"].alignment, dm["B2"].border = (
    font(10, True), CALC_FILL, CTR, BOX)
header_row(dm, 4, [lab for _, lab, _ in DM_COLS], start_col=1, height=30)
DM_R0 = 5
for i in range(N_DOC):
    r = DM_R0 + i; sr = DS_ROW0 + i
    rng = f"醫師班表!${E0}{sr}:${EL}{sr}"
    g = f'IF($A{r}="","",'
    vals = {
        "A": f'=IF(醫師班表!$A{sr}="","",醫師班表!$A{sr})',
        "B": f'={g}醫師班表!$B{sr})',
        "C": f'={g}醫師班表!$C{sr})',
        "D": f'={g}醫師班表!$D{sr})',
        "J": f'={g}SUM($E{r}:$I{r}))',
        "K": f'={g}SUMPRODUCT(($E{r}:$I{r}>0)*1))',
        "L": f'={g}COUNTIF({rng},"訓"))',
        "T": f'={g}SUM($M{r}:$S{r}))',
        "U": f'={g}{DAYS_FX}*3-COUNTA({rng}))',
    }
    for k, c in enumerate(CLINIC_CODES):
        vals[get_column_letter(5+k)] = f'={g}COUNTIF({rng},"{c}"))'
    for k, code in enumerate(DOC_LEAVE):
        vals[get_column_letter(13+k)] = f'={g}COUNTIF({rng},"{code}"))'
    for col, _, _w in DM_COLS:
        c = dm[f"{col}{r}"]
        c.value, c.font, c.border, c.alignment, c.fill = (
            vals[col], font(9), BOX, CTR, CALC_FILL)
        if col == "C": c.alignment = LEFT
        if col in [get_column_letter(5+k) for k in range(5)]:
            c.fill = CLINIC_FILL[CLINIC_CODES[[get_column_letter(5+k)
                     for k in range(5)].index(col)]]
DM_TOT = DM_R0 + N_DOC
put(dm, f"A{DM_TOT}", "合計", font(10, True), SUB_FILL, CTR)
dm.merge_cells(f"A{DM_TOT}:D{DM_TOT}")
for k in range(4, 21):
    col = get_column_letter(k+1)
    c = dm[f"{col}{DM_TOT}"]
    c.value = f"=SUM({col}{DM_R0}:{col}{DM_TOT-1})"
    c.font, c.fill, c.border, c.alignment = font(10, True), SUB_FILL, BOX, CTR
for i, t in enumerate([
  "※ 醫師以「診次」為單位,不是天數。請假 3 個診次等於請一天。",
  "※ 服務院所數 = 本月實際有排診的院所家數,可用來看跨院負荷。",
  "※ 未排診次 = 當月總診次格數(天數 × 3)扣掉已填格數,不代表應該排滿。",
]):
    put(dm, f"A{DM_TOT+2+i}", t, font(9, color="808080"), None, LEFT, border=False)


# ================================================================ 5. 助理班表
AS_C0, AS_C1 = 5, 5 + DAYS_IN_MONTH - 1      # E..AI
AS_ROW0 = 5
AS_ROW1 = AS_ROW0 + N_ASST - 1               # 64
A0 = get_column_letter(AS_C0); A1 = get_column_letter(AS_C1)

asx = wb.create_sheet("助理班表")
asx.sheet_view.showGridLines = False
asx.freeze_panes = "E5"
for col, w in {"A":9,"B":10,"C":9,"D":10}.items():
    asx.column_dimensions[col].width = w
for c in range(AS_C0, AS_C1 + 1):
    asx.column_dimensions[get_column_letter(c)].width = 4.4
put(asx, "A1", "助理 / 醫護長 班表(工時制)", TITLE_F, border=False)
asx.merge_cells(start_row=1, start_column=1, end_row=1, end_column=AS_C1)
put(asx, "A2", "期間", font(10, True), SUB_FILL, CTR)
asx["B2"] = f'={SET_Y}&"年"&{SET_M}&"月"'
asx["B2"].font, asx["B2"].fill, asx["B2"].alignment, asx["B2"].border = (
    font(10, True), CALC_FILL, CTR, BOX)
put(asx, "E2", "← 各院所醫護長只填自己院所那幾列。院所欄由人員名冊自動帶出。",
    font(9, color="808080"), None, LEFT, border=False)

for lab, col in (("員工編號",1), ("姓名",2), ("職類",3), ("院所",4)):
    c = asx.cell(row=3, column=col, value=lab)
    c.font, c.fill, c.alignment, c.border = HDR_F, HDR_FILL, CTR, BOX
    asx.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)
for c in range(AS_C0, AS_C1 + 1):
    d = c - AS_C0 + 1; L = get_column_letter(c)
    h = asx.cell(row=3, column=c, value=f'=IF({d}>{DAYS_FX},"",{d})')
    h.font, h.fill, h.alignment, h.border = HDR_F, HDR_FILL, CTR, BOX
    w = asx.cell(row=4, column=c)
    w.value = (f'=IF({L}$3="","",'
               f'INDEX({R_WEEK},WEEKDAY(DATE({SET_Y},{SET_M},{L}$3),1)))')
    w.font, w.fill, w.alignment, w.border = font(9, True), SUB_FILL, CTR, BOX

for i in range(N_ASST):
    r = AS_ROW0 + i
    asx.row_dimensions[r].height = 17
    eid = ASSISTANTS[i][0] if i < len(ASSISTANTS) else None
    a = asx.cell(row=r, column=1, value=eid)
    a.font, a.fill, a.border, a.alignment = font(9), IN_FILL, BOX, CTR
    for col, src in ((2, R_NAME), (3, R_ROLE), (4, R_HOME)):
        c = asx.cell(row=r, column=col)
        c.value = f'=IFERROR(INDEX({src},MATCH($A{r},{R_EID},0)),"")'
        c.font, c.fill, c.border, c.alignment = font(9), CALC_FILL, BOX, CTR
    for c in range(AS_C0, AS_C1 + 1):
        cc = asx.cell(row=r, column=c)
        cc.font, cc.border, cc.alignment = font(9), BOX, CTR

AS_TALLY = AS_ROW1 + 2
put(asx, f"A{AS_TALLY-1}", "各院所當日在班人數(自動計算)", font(10, True), SUB_FILL, LEFT)
asx.merge_cells(start_row=AS_TALLY-1, start_column=1, end_row=AS_TALLY-1, end_column=4)
for k, (code, short, *_r) in enumerate(CLINICS):
    r = AS_TALLY + k
    put(asx, f"A{r}", f"{code} {short}", font(9, True), CLINIC_FILL[code], CTR)
    asx.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    for c in range(AS_C0, AS_C1 + 1):
        L = get_column_letter(c)
        cell = asx.cell(row=r, column=c)
        cell.value = (f'=IF({L}$3="","",SUMPRODUCT(($D${AS_ROW0}:$D${AS_ROW1}="{short}")'
                      f'*(COUNTIF({R_WC_W},{L}${AS_ROW0}:{L}${AS_ROW1})>0)))')
        cell.font, cell.fill, cell.alignment, cell.border = (
            font(9, True), CALC_FILL, CTR, BOX)
        cell.number_format = FMT_CNT
put(asx, f"A{AS_TALLY+len(CLINICS)+1}",
    "※ 門檻:有看診的日子每間院所至少 2 位在班,低於門檻自動變紅;全院休診日不示警。"
    "門檻要調請改這幾列的條件式格式。", font(9, color="808080"), None, LEFT, border=False)

dv_wc = DataValidation(type="list", formula1=R_WC, allow_blank=True,
                       showErrorMessage=True, errorTitle="班別代碼無效",
                       error="請從下拉選單選擇「設定」分頁定義的班別代碼。")
asx.add_data_validation(dv_wc)
dv_wc.add(f"{A0}{AS_ROW0}:{A1}{AS_ROW1}")
dv_aeid = DataValidation(type="list", formula1=f"設定!$S$6:$S${5+N_ASST_SLOTS}",
                         allow_blank=True)
asx.add_data_validation(dv_aeid); dv_aeid.add(f"A{AS_ROW0}:A{AS_ROW1}")

AGRID = f"{A0}{AS_ROW0}:{A1}{AS_ROW1}"
asx.conditional_formatting.add(AGRID, FormulaRule(
    formula=[f'AND({A0}$3<>"",$A{AS_ROW0}<>"",{A0}{AS_ROW0}="")'],
    fill=GAP_FILL, stopIfTrue=True))
asx.conditional_formatting.add(AGRID, FormulaRule(
    formula=[f'{A0}{AS_ROW0}="支"'], fill=CLINIC_FILL["睿"], stopIfTrue=True))
asx.conditional_formatting.add(AGRID, FormulaRule(
    formula=[f'COUNTIF({R_WC_L},{A0}{AS_ROW0})>0'], fill=LEAVE_FILL, stopIfTrue=True))
asx.conditional_formatting.add(AGRID, FormulaRule(
    formula=[f'OR({A0}$4="六",{A0}$4="日")'], fill=WKND_FILL))
asx.conditional_formatting.add(f"{A0}3:{A1}4", FormulaRule(
    formula=[f'OR({A0}$4="六",{A0}$4="日")'], fill=WKND_FILL))
for k in range(len(CLINICS)):
    r = AS_TALLY + k
    asx.conditional_formatting.add(f"{A0}{r}:{A1}{r}", FormulaRule(
        formula=[f'AND({A0}$3<>"",{A0}{r}>0,{A0}{r}<2)'], fill=ALERT_FILL))

# ================================================================ 6. 打卡匯入
PUNCH_R0 = 3
PUNCH_R1 = PUNCH_R0 + PUNCH_N - 1
pc = wb.create_sheet("打卡匯入")
pc.sheet_view.showGridLines = False
pc.freeze_panes = "A3"
for col, w in {"A":13,"B":13,"C":12,"D":12,"E":16,"F":16,"G":2,"H":76}.items():
    pc.column_dimensions[col].width = w
put(pc, "A1", "打卡匯入(把打卡機匯出的資料貼在下面四欄)", TITLE_F, border=False)
pc.merge_cells("A1:F1")
put(pc, "H1", "貼上規則:一天一列。日期要是真正的日期格式、時間要是真正的時間格式,"
    "不能是文字。E、F 欄是公式,不要覆蓋。", font(9, color="808080"), None, WRAP, border=False)
header_row(pc, 2, ["員工編號","日期","上班時間","下班時間","對照鍵(自動)","檢核(自動)"],
           start_col=1, height=24)
for i in range(PUNCH_R0, PUNCH_R1 + 1):
    for j in range(4):
        c = pc.cell(row=i, column=1 + j)
        c.font, c.border, c.alignment, c.fill = font(9), BOX, CTR, IN_FILL
        if j == 1: c.number_format = FMT_DATE
        if j in (2, 3): c.number_format = FMT_TIME
    e = pc.cell(row=i, column=5)
    e.value = f'=IF($A{i}="","",$A{i}&"|"&DAY($B{i}))'
    e.font, e.border, e.alignment, e.fill = font(9), BOX, CTR, CALC_FILL
    f = pc.cell(row=i, column=6)
    f.value = (f'=IF($A{i}="","",'
               f'IF(COUNTIF({R_EID},$A{i})=0,"⚠ 查無員編",'
               f'IF(COUNTIF($E${PUNCH_R0}:$E${PUNCH_R1},$E{i})>1,"⚠ 重複打卡","OK")))')
    f.font, f.border, f.alignment, f.fill = font(9), BOX, CTR, CALC_FILL
pc.auto_filter.ref = f"A2:F{PUNCH_R1}"
pc.conditional_formatting.add(f"A{PUNCH_R0}:F{PUNCH_R1}", FormulaRule(
    formula=[f'AND($A{PUNCH_R0}<>"",LEFT($F{PUNCH_R0},1)="⚠")'], fill=GAP_FILL))

# ================================================================ 7. 出勤紀錄
ATT_R0 = 3
ATT_R1 = ATT_R0 + N_ASST * DAYS_IN_MONTH - 1
at = wb.create_sheet("出勤紀錄")
at.sheet_view.showGridLines = False
at.freeze_panes = "E3"
AT_COLS = [("A","日期",11), ("B","星期",6), ("C","員工編號",11), ("D","姓名",11),
           ("E","院所",10), ("F","職類",9), ("G","排班代碼",9), ("H","應到",8),
           ("I","應退",8), ("J","實際上班",10), ("K","實際下班",10),
           ("L","休息(分)",9), ("M","實際工時",10), ("N","出勤異常",11),
           ("O","遲到(分)",9), ("P","早退(分)",9), ("Q","加班時數",10),
           ("R","加班別",10), ("S","備註",22)]
for col, _, w in AT_COLS: at.column_dimensions[col].width = w
put(at, "A1", "出勤紀錄(法定紀錄:逐日記載至分鐘,保存 5 年)", TITLE_F, border=False)
at.merge_cells("A1:S1")
header_row(at, 2, [lab for _, lab, _ in AT_COLS], start_col=1, height=28)
for pi in range(N_ASST):
    srow = AS_ROW0 + pi
    for d in range(1, DAYS_IN_MONTH + 1):
        r = ATT_R0 + pi * DAYS_IN_MONTH + (d - 1)
        scol = get_column_letter(AS_C0 + d - 1)
        DUE = f'IFERROR(INDEX({R_WC_HRS},MATCH($G{r},{R_WC},0)),0)'
        fx = {
"A": f'=IF(OR(助理班表!$A{srow}="",{d}>{DAYS_FX}),"",DATE({SET_Y},{SET_M},{d}))',
"B": f'=IF($A{r}="","",INDEX({R_WEEK},WEEKDAY($A{r},1)))',
"C": f'=IF($A{r}="","",助理班表!$A{srow})',
"D": f'=IF($C{r}="","",助理班表!$B{srow})',
"E": f'=IF($C{r}="","",助理班表!$D{srow})',
"F": f'=IF($C{r}="","",助理班表!$C{srow})',
"G": f'=IF($C{r}="","",助理班表!{scol}{srow})',
"H": f'=IF($G{r}="","",IFERROR(INDEX({R_WC_IN},MATCH($G{r},{R_WC},0)),""))',
"I": f'=IF($G{r}="","",IFERROR(INDEX({R_WC_OUT},MATCH($G{r},{R_WC},0)),""))',
"J": (f'=IFERROR(INDEX(打卡匯入!$C${PUNCH_R0}:$C${PUNCH_R1},'
      f'MATCH($C{r}&"|"&DAY($A{r}),打卡匯入!$E${PUNCH_R0}:$E${PUNCH_R1},0)),"")'),
"K": (f'=IFERROR(INDEX(打卡匯入!$D${PUNCH_R0}:$D${PUNCH_R1},'
      f'MATCH($C{r}&"|"&DAY($A{r}),打卡匯入!$E${PUNCH_R0}:$E${PUNCH_R1},0)),"")'),
"L": f'=IF(OR($J{r}="",$K{r}=""),"",IFERROR(INDEX({R_WC_RST},MATCH($G{r},{R_WC},0)),0))',
"M": f'=IF(OR($J{r}="",$K{r}=""),"",ROUND(($K{r}-$J{r})*24-$L{r}/60,2))',
"O": (f'=IF(OR($C{r}="",$H{r}="",$J{r}=""),0,'
      f'IF(ROUND(($J{r}-$H{r})*1440,0)>{P_GRACE},ROUND(($J{r}-$H{r})*1440,0),0))'),
"P": (f'=IF(OR($C{r}="",$I{r}="",$K{r}=""),0,'
      f'IF(ROUND(($I{r}-$K{r})*1440,0)>{P_GRACE},ROUND(($I{r}-$K{r})*1440,0),0))'),
"N": (f'=IF($C{r}="","",'
      f'IF(AND(COUNTIF({R_WC_L},$G{r})>0,$J{r}<>""),"假日出勤",'
      f'IF(AND(COUNTIF({R_WC_W},$G{r})>0,$J{r}=""),"未打卡",'
      f'IF($J{r}="","",IF($O{r}>0,"遲到",IF($P{r}>0,"早退","正常"))))))'),
"Q": (f'=IF($M{r}="",0,IF(($M{r}-{DUE})*60>={P_OT_MIN},'
      f'ROUND(FLOOR(($M{r}-{DUE})*60,{P_OT_UNIT})/60,2),0))'),
"R": (f'=IF(OR($C{r}="",$Q{r}=0),"",IF($G{r}="國","國定假日",'
      f'IF($B{r}="日","例假",IF($B{r}="六","休息日","平日"))))'),
"S": None,
        }
        for col, _, _w in AT_COLS:
            c = at[f"{col}{r}"]
            v = fx.get(col)
            if v is not None: c.value = v; c.fill = CALC_FILL
            else: c.fill = IN_FILL
            c.font, c.border, c.alignment = font(9), BOX, CTR
            if col == "A": c.number_format = FMT_DATE
            elif col in ("H","I","J","K"): c.number_format = FMT_TIME
            elif col in ("M","Q"): c.number_format = FMT_HR
            elif col in ("L","O","P"): c.number_format = FMT_MIN
            elif col == "S": c.alignment = LEFT
at.auto_filter.ref = f"A2:S{ATT_R1}"
AT_RANGE = f"A{ATT_R0}:S{ATT_R1}"
at.conditional_formatting.add(AT_RANGE, FormulaRule(
    formula=[f'AND($C{ATT_R0}<>"",$N{ATT_R0}<>"",$N{ATT_R0}<>"正常")'],
    fill=GAP_FILL, stopIfTrue=True))
at.conditional_formatting.add(AT_RANGE, FormulaRule(
    formula=[f'AND($C{ATT_R0}<>"",$Q{ATT_R0}>0)'], fill=OT_FILL))
at.conditional_formatting.add(f"N{ATT_R0}:N{ATT_R1}", FormulaRule(
    formula=[f'AND($N{ATT_R0}<>"",$N{ATT_R0}<>"正常")'], fill=ALERT_FILL))

# ================================================================ 8. 月結統計
ms = wb.create_sheet("月結統計")
ms.sheet_view.showGridLines = False
ms.freeze_panes = "E5"
MS_COLS = [("A","員工編號",11), ("B","姓名",11), ("C","職類",9), ("D","院所",10),
           ("E","出勤天數",9), ("F","排班工時",9), ("G","排休",7), ("H","特休",7),
           ("I","病假",7), ("J","事假",7), ("K","公假",7), ("L","國定假日",9),
           ("M","支援他院",9), ("N","加班時數",10), ("O","實際工時",10),
           ("P","遲到次數",9), ("Q","早退次數",9), ("R","未打卡",8),
           ("S","排班完整度檢核",15)]
for col, _, w in MS_COLS: ms.column_dimensions[col].width = w
put(ms, "A1", "月結統計 — 助理 / 醫護長(全自動)", TITLE_F, border=False)
ms.merge_cells("A1:S1")
put(ms, "A2", "期間", font(10, True), SUB_FILL, CTR)
ms["B2"] = f'={SET_Y}&"年"&{SET_M}&"月"'
ms["B2"].font, ms["B2"].fill, ms["B2"].alignment, ms["B2"].border = (
    font(10, True), CALC_FILL, CTR, BOX)
header_row(ms, 4, [lab for _, lab, _ in MS_COLS], start_col=1, height=30)
MS_R0 = 5
AT_C = f"出勤紀錄!$C${ATT_R0}:$C${ATT_R1}"
AT_N = f"出勤紀錄!$N${ATT_R0}:$N${ATT_R1}"
AT_Q = f"出勤紀錄!$Q${ATT_R0}:$Q${ATT_R1}"
for i in range(N_ASST):
    r = MS_R0 + i; sr = AS_ROW0 + i
    rng = f"助理班表!${A0}{sr}:${A1}{sr}"
    g = f'IF($A{r}="","",'
    vals = {
"A": f'=IF(助理班表!$A{sr}="","",助理班表!$A{sr})',
"B": f'={g}助理班表!$B{sr})',
"C": f'={g}助理班表!$C{sr})',
"D": f'={g}助理班表!$D{sr})',
"E": f'={g}SUMPRODUCT(COUNTIF({rng},{R_WC}),{R_WC_ATT}))',
"F": f'={g}SUMPRODUCT(COUNTIF({rng},{R_WC}),{R_WC_HRS}))',
"G": f'={g}COUNTIF({rng},"OFF"))',
"H": f'={g}COUNTIF({rng},"特"))',
"I": f'={g}COUNTIF({rng},"病"))',
"J": f'={g}COUNTIF({rng},"事"))',
"K": f'={g}COUNTIF({rng},"公"))',
"L": f'={g}COUNTIF({rng},"國"))',
"M": f'={g}COUNTIF({rng},"支"))',
"N": f'={g}SUMIF({AT_C},$A{r},{AT_Q}))',
"O": f'={g}$F{r}+$N{r})',
"P": f'={g}COUNTIFS({AT_C},$A{r},{AT_N},"遲到"))',
"Q": f'={g}COUNTIFS({AT_C},$A{r},{AT_N},"早退"))',
"R": f'={g}COUNTIFS({AT_C},$A{r},{AT_N},"未打卡"))',
"S": (f'={g}IF(COUNTA({rng})={DAYS_FX},"OK",'
      f'"⚠ 缺"&({DAYS_FX}-COUNTA({rng}))&"天"))'),
    }
    for col, _, _w in MS_COLS:
        c = ms[f"{col}{r}"]
        c.value, c.font, c.border, c.alignment, c.fill = (
            vals[col], font(9), BOX, CTR, CALC_FILL)
        if col in ("F","N","O"): c.number_format = "0.0"
MS_TOT = MS_R0 + N_ASST
put(ms, f"A{MS_TOT}", "合計", font(10, True), SUB_FILL, CTR)
ms.merge_cells(f"A{MS_TOT}:D{MS_TOT}")
for col in "EFGHIJKLMNOPQR":
    c = ms[f"{col}{MS_TOT}"]
    c.value = f"=SUM({col}{MS_R0}:{col}{MS_TOT-1})"
    c.font, c.fill, c.border, c.alignment = font(10, True), SUB_FILL, BOX, CTR
    if col in ("F","N","O"): c.number_format = "0.0"
put(ms, f"S{MS_TOT}", "", font(), SUB_FILL, CTR)
ms.conditional_formatting.add(f"S{MS_R0}:S{MS_TOT-1}", FormulaRule(
    formula=[f'AND($A{MS_R0}<>"",S{MS_R0}<>"OK")'], fill=ALERT_FILL))
for col in ("P","Q","R"):
    ms.conditional_formatting.add(f"{col}{MS_R0}:{col}{MS_TOT-1}", FormulaRule(
        formula=[f'AND($A{MS_R0}<>"",{col}{MS_R0}>0)'], fill=GAP_FILL))
for i, t in enumerate([
  "※ 出勤天數與排班工時依「設定」分頁的班別代碼權重自動加總,改權重這裡跟著變。",
  "※ 加班時數由「出勤紀錄」加總,已套用加班門檻。不使用打卡匯入的院所可覆蓋成手動數字。",
  "※ 加班時數為「實際工時 − 排班工時」,未依勞基法第 24 條換算費率,薪資請另行計算。",
  "※ 醫師的月結在「醫師月結」分頁,以診次為單位計算。",
]):
    put(ms, f"A{MS_TOT+2+i}", t, font(9, color="808080"), None, LEFT, border=False)

wb.save(OUT)
print("saved:", OUT)
print(f"醫師班表:{N_DOC} 列 × {DAYS_IN_MONTH*3} 診次欄")
print(f"助理班表:{N_ASST} 列 · 出勤紀錄:{N_ASST*DAYS_IN_MONTH} 列 · 打卡匯入:{PUNCH_N} 列")
print("分頁:", wb.sheetnames)
